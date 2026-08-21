import csv
import io
import os
import uuid
from datetime import UTC, datetime
from typing import Any

from pawguard.core.config import get_settings


def generate_csv(
    headers: list[str],
    rows: list[list[Any]],
    sections: list[dict] | None = None,
) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        cleaned = [str(cell) if cell is not None else "" for cell in row]
        writer.writerow(cleaned)
    if sections:
        for section in sections:
            writer.writerow([])
            writer.writerow([section.get("title", "")])
            writer.writerow(section.get("headers", []))
            for row in section.get("rows", []):
                writer.writerow([str(cell) if cell is not None else "" for cell in row])
    return output.getvalue().encode("utf-8-sig")


def generate_excel(
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    title: str | None = None,
    sections: list[dict] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)

    header_row = 2 if title else 1
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    data_start = header_row + 1
    for row_idx, row in enumerate(rows, start=data_start):
        for col_idx, cell_value in enumerate(row, start=1):
            val = str(cell_value) if cell_value is not None else ""
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(header_row, data_start + len(rows))
        )
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 4, 50)

    # Secondary sections render as their own worksheets (e.g. campaign
    # progress metrics inside the Donor & Financial Reconciliation Report).
    if sections:
        for section in sections:
            sec_ws = wb.create_sheet(title=f"{section.get('title', 'Section')}"[:31])
            sec_headers = section.get("headers", [])
            sec_rows = section.get("rows", [])
            for col_idx, header in enumerate(sec_headers, start=1):
                cell = sec_ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
            for row_idx, row in enumerate(sec_rows, start=2):
                for col_idx, cell_value in enumerate(row, start=1):
                    sec_ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=str(cell_value) if cell_value is not None else "",
                    )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    subtitle: str | None = None,
    landscape: bool = False,
    sections: list[dict] | None = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.pagesizes import landscape as landscape_size
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    page_size = landscape_size(A4) if landscape else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Title"]))
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [headers]
    for row in rows:
        table_data.append([str(c) if c is not None else "" for c in row])

    available_width = page_size[0] - inch
    col_width = available_width / max(len(headers), 1)
    table = Table(table_data, colWidths=[col_width] * len(headers))

    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]
    )
    table.setStyle(style)
    elements.append(table)

    # Secondary sections (e.g. campaign progress metrics) render below the
    # main table as their own titled tables.
    if sections:
        for section in sections:
            elements.append(Spacer(1, 14))
            elements.append(Paragraph(section.get("title", "Section"), styles["Heading2"]))
            elements.append(Spacer(1, 6))
            sec_headers = section.get("headers", [])
            sec_rows = section.get("rows", [])
            sec_data = [sec_headers]
            sec_data.extend([str(c) if c is not None else "" for c in row] for row in sec_rows)
            if sec_headers:
                sec_width = available_width / len(sec_headers)
                sec_table = Table(sec_data, colWidths=[sec_width] * len(sec_headers))
                sec_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 9),
                            ("FONTSIZE", (0, 1), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            (
                                "ROWBACKGROUNDS",
                                (0, 1),
                                (-1, -1),
                                [colors.white, colors.HexColor("#F2F2F2")],
                            ),
                        ]
                    )
                )
                elements.append(sec_table)

    doc.build(elements)
    return buf.getvalue()


def report_filename(report_type: str, fmt: str) -> str:
    today = datetime.now(UTC).strftime("%Y%m%d")
    rid = uuid.uuid4().hex[:8]
    return f"{report_type}_{today}_{rid}.{fmt}"


def ensure_reports_dir() -> str:
    settings = get_settings()
    reports_dir = getattr(settings, "reports_dir", None) or os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "generated_reports"
    )
    os.makedirs(reports_dir, exist_ok=True)
    return reports_dir
